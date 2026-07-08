"""Strategy 1 -- build the formatted Excel workbook from the backtest outputs.

Reads strategy-1-bet-log.csv and strategy-1-weekly-pnl.csv (produced by
``scripts/strategy-1.py``) and writes strategy-1-epl-backtest.xlsx with three
tabs: Summary (rules + headline + by-category), Bet Log (all bets with running
bankroll/drawdown, auto-filter, red/green profit), and Weekly P&L (per-week
ledger + a bankroll line chart).

Values are computed with pandas and written directly (not as live formulas) so
the workbook opens error-free in any viewer without a recalculation engine.

    python scripts/strategy-1-report.py
"""
from __future__ import annotations

import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from soccer_backtest import config  # noqa: E402

FONT = "Arial"
navy, blue, grey = "1F3864", "2A78D6", "F2F2F2"
hdr_fill = PatternFill("solid", fgColor=navy)
band = PatternFill("solid", fgColor=grey)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '$#,##0.00;($#,##0.00)'
MONEY0 = '$#,##0;($#,##0)'
PCT, PCT2, ODDS = '0.0%', '0.00%', '0.00'
red = Font(name=FONT, size=10, bold=True, color="C00000")


def header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def main() -> None:
    d = config.PROCESSED_DIR
    bl = pd.read_csv(d / "strategy-1-bet-log.csv")
    wk = pd.read_csv(d / "strategy-1-weekly-pnl.csv")
    bl["bankroll"] = 1000 + bl["profit"].cumsum()
    bl["peak"] = bl["bankroll"].cummax()
    bl["dd"] = bl["bankroll"] - bl["peak"]
    bl["ddpct"] = bl["bankroll"] / bl["peak"] - 1

    wb = Workbook()

    # ---------------- Bet Log ----------------
    ws = wb.create_sheet("Bet Log")
    cols = ["Week", "Date", "Season", "Home", "Away", "Score", "Category",
            "Selection", "Est. prob", "Odds", "Stake ($)", "Result", "Profit ($)",
            "Bankroll ($)", "Peak ($)", "Drawdown ($)", "Drawdown %"]
    ws.append(cols); header(ws, 1, len(cols))
    last = len(bl) + 1
    for i, r in enumerate(bl.itertuples(index=False), start=2):
        vals = [r.week, r.date, r.season, r.home, r.away, str(r.score), r.category,
                str(r.selection), round(float(r.est_prob), 3), round(float(r.odds), 2),
                1, r.result, round(float(r.profit), 2), round(float(r.bankroll), 2),
                round(float(r.peak), 2), round(float(r.dd), 2), round(float(r.ddpct), 4)]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v); cell.font = Font(name=FONT, size=10); cell.border = border
        ws.cell(i, 9).number_format = PCT
        ws.cell(i, 10).number_format = ODDS
        ws.cell(i, 11).number_format = MONEY0
        for c in (13, 14, 15, 16):
            ws.cell(i, c).number_format = MONEY
        ws.cell(i, 17).number_format = PCT2
        for c in (7, 8, 12):
            ws.cell(i, c).alignment = Alignment(horizontal="center")
    for c, w in enumerate([10, 11, 10, 16, 16, 7, 8, 8, 9, 7, 9, 7, 11, 12, 11, 12, 11], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:Q{last}"
    for op, col in [("lessThan", "C00000"), ("greaterThan", "1B7A3D")]:
        ws.conditional_formatting.add(f"M2:M{last}",
            CellIsRule(operator=op, formula=["0"], font=Font(name=FONT, size=10, color=col)))

    # ---------------- Weekly P&L ----------------
    wsw = wb.create_sheet("Weekly P&L")
    wcols = ["Week", "Season", "Games", "Bets", "Staked ($)", "P&L ($)",
             "Cumulative P&L ($)", "Bankroll ($)"]
    wsw.append(wcols); header(wsw, 1, len(wcols))
    wlast = len(wk) + 1
    for i, r in enumerate(wk.itertuples(index=False), start=2):
        seas = "2025-2026" if r.week >= "2025-W30" else "2024-2025"
        vals = [r.week, seas, int(r.games), int(r.bets), round(float(r.staked), 2),
                round(float(r.pnl), 2), round(float(r.cum_pnl), 2), round(float(r.bankroll), 2)]
        for c, v in enumerate(vals, 1):
            cell = wsw.cell(i, c, v); cell.font = Font(name=FONT, size=10)
            cell.border = border; cell.alignment = Alignment(horizontal="center")
        for c in (5, 6, 7, 8):
            wsw.cell(i, c).number_format = MONEY
        if i % 2 == 0:
            for c in range(1, len(wcols) + 1):
                wsw.cell(i, c).fill = band
    for c, w in zip(range(1, 9), [10, 11, 8, 7, 11, 11, 16, 13]):
        wsw.column_dimensions[get_column_letter(c)].width = w
    wsw.freeze_panes = "A2"
    for op, col in [("lessThan", "C00000"), ("greaterThan", "1B7A3D")]:
        wsw.conditional_formatting.add(f"F2:F{wlast}",
            CellIsRule(operator=op, formula=["0"], font=Font(name=FONT, size=10, color=col)))

    chart = LineChart(); chart.title = "Account bankroll by week ($1,000 start)"
    chart.style = 2; chart.height = 8.5; chart.width = 22
    chart.y_axis.title = "Bankroll ($)"; chart.x_axis.title = "Week"
    chart.y_axis.scaling.min = 850; chart.y_axis.scaling.max = 1020
    chart.x_axis.delete = False; chart.y_axis.delete = False
    chart.add_data(Reference(wsw, min_col=8, min_row=1, max_row=wlast), titles_from_data=True)
    chart.set_categories(Reference(wsw, min_col=1, min_row=2, max_row=wlast))
    chart.series[0].graphicalProperties.line.solidFill = blue
    chart.series[0].graphicalProperties.line.width = 22000
    chart.legend = None
    wsw.add_chart(chart, "J2")

    # ---------------- Summary ----------------
    wss = wb.active; wss.title = "Summary"
    wss.sheet_view.showGridLines = False

    def put(row, col, v, bold=False, size=10, color="000000", italic=False,
            fmt=None, fill=None, align=None):
        cell = wss.cell(row, col, v)
        cell.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
        if fmt: cell.number_format = fmt
        if fill: cell.fill = PatternFill("solid", fgColor=fill)
        if align: cell.alignment = Alignment(horizontal=align)
        return cell

    is1 = bl.category == "1X2"
    tot_stk, tot_pnl = float(bl.stake.sum()), float(bl.profit.sum())

    put(1, 1, "Strategy 1 -- favourite / total-goals backtest", bold=True, size=16, color=navy)
    put(2, 1, "English Premier League - 2024-25 + 2025-26 - $1,000 starting capital",
        italic=True, color="595959")
    r = 4
    put(r, 1, "Strategy rules", bold=True, size=11, color="FFFFFF", fill=navy)
    for cc in (2, 3): put(r, cc, "", fill=navy)
    for t in [
        "Probabilities are de-vigged implied odds:  p = (1/odds) / sum(1/odds).",
        "1X2: stake $1 on the single most-likely outcome only if its prob > 70%.",
        "Total goals (0..9, where 9 = 9 or more): stake $1 on each of the 2 most",
        "     probable outcomes.",
        f"Selection: each ISO week, bet the highest-confidence games (up to "
        f"{int(wk.games.max())} bet in a week here).",
        "Flat $1 per bet; bets settle at the same odds used to estimate probability.",
    ]:
        r += 1; put(r, 1, t)
    r += 1
    put(r, 1, "Data:  1X2 = football-data Bet365 closing (both seasons).  Total goals = "
              "Singapore Pools / sgodds, 2025-26 only (no exact-TG data exists for 2024-25).",
        size=9, italic=True, color="595959")

    r += 2
    put(r, 1, "Headline", bold=True, color="FFFFFF", fill=navy)
    put(r, 2, "Value", bold=True, color="FFFFFF", fill=navy, align="right")
    for lbl, v, fmt, danger in [
        ("Starting capital", 1000, MONEY0, False),
        ("Total bets placed", len(bl), '#,##0', False),
        ("   of which 1X2", int(is1.sum()), '#,##0', False),
        ("   of which total-goals legs", int((~is1).sum()), '#,##0', False),
        ("Total staked", tot_stk, MONEY0, False),
        ("Total profit / loss", tot_pnl, MONEY, True),
        ("Return on stake (ROI)", tot_pnl / tot_stk, PCT, True),
        ("Final bankroll", 1000 + tot_pnl, MONEY, False),
        ("Maximum drawdown ($)", float(bl.dd.min()), MONEY, True),
        ("Maximum drawdown (%)", float(bl.ddpct.min()), PCT2, True),
        ("1X2 hit rate", float(bl[is1].result.eq("W").mean()), PCT, False),
    ]:
        r += 1
        put(r, 1, lbl)
        cell = put(r, 2, v, fmt=fmt, align="right")
        if danger: cell.font = red

    r += 2
    put(r, 1, "By category", bold=True, color="FFFFFF", fill=navy)
    for c, t in zip(range(2, 7), ["Bets", "Staked", "P&L", "ROI", "Hit rate"]):
        put(r, c, t, bold=True, color="FFFFFF", fill=navy, align="center")
    for lbl, mask in [("1X2 (favourites)", is1), ("Total goals (2 outcomes)", ~is1)]:
        sub = bl[mask]
        stk, pnl = float(sub.stake.sum()), float(sub.profit.sum())
        r += 1
        put(r, 1, lbl)
        put(r, 2, len(sub), fmt='#,##0', align="right")
        put(r, 3, stk, fmt=MONEY0, align="right")
        pc = put(r, 4, pnl, fmt=MONEY, align="right")
        if pnl < 0: pc.font = red
        put(r, 5, pnl / stk, fmt=PCT, align="right")
        put(r, 6, float(sub.result.eq("W").mean()), fmt=PCT, align="right")

    for cc, w in zip(range(1, 7), [30, 12, 12, 12, 10, 10]):
        wss.column_dimensions[get_column_letter(cc)].width = w

    out = d / "strategy-1-epl-backtest.xlsx"
    wb.save(out)
    print(f"wrote {out}  ({len(bl)} bets, {len(wk)} weeks, "
          f"P&L {tot_pnl:+.1f}, maxDD {float(bl.dd.min()):+.1f})")


if __name__ == "__main__":
    main()
